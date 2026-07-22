from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
ARCHIVE_CSV = ROOT / "data" / "paper_archive" / "stage2_bestq_original_paper.csv"
NPZ_DIR = ROOT / "output" / "npz"
OUT_CSV = ROOT / "output" / "csv" / "stage2_bestq_original_with_dqn.csv"
OUT_JSON = ROOT / "output" / "csv" / "stage2_bestq_original_with_dqn.json"
OUT_VALIDATION = ROOT / "output" / "csv" / "stage2_bestq_original_with_dqn_validation.csv"

METHOD_ORDER = ["NS-P", "PSP", "GCP", "GDP", "DQN"]
CONFIG_ORDER = ["10_100", "10_130", "10_150", "10_180", "5_130", "15_130", "20_130"]
AUDIT_METHODS = ["nsp", "gcp", "gdp", "psp"]


def repo_path(path: Path) -> str:
    """Return a stable repository-relative path for exported provenance fields."""
    return path.resolve().relative_to(ROOT.resolve()).as_posix()


def objective_bounds_from_workbook(path: Path) -> tuple[np.ndarray, np.ndarray]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    points: list[tuple[float, float]] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            cost, delay = row[0], row[1]
            if isinstance(cost, (int, float)) and isinstance(delay, (int, float)):
                points.append((float(cost), float(delay)))
    if not points:
        raise ValueError(f"No objective values found in {path}")
    objectives = np.asarray(points, dtype=float)
    return objectives.min(axis=0), objectives.max(axis=0)


def objective_bounds_from_npz(paths: list[Path]) -> tuple[np.ndarray, np.ndarray]:
    objectives = []
    for path in paths:
        if not path.exists():
            raise FileNotFoundError(path)
        with np.load(path) as data:
            objectives.append(np.asarray(data["F"], dtype=float))
    combined = np.vstack(objectives)
    return combined.min(axis=0), combined.max(axis=0)


def audit_paths(config: str) -> list[Path]:
    folder = NPZ_DIR / "paper_protocol_audit" / "seed42" / "reset_per_method"
    suffix = "configured"
    if config == "10_150":
        suffix = "new"
    elif config == "10_180":
        suffix = "new"
    return [folder / f"{config}_{suffix}_{method}.npz" for method in AUDIT_METHODS]


def normalization_bounds(config: str) -> tuple[np.ndarray, np.ndarray, str, str]:
    if config == "5_130":
        path = ROOT / "output" / "excel" / "nsga2_normalized_results.xlsx"
        lower, upper = objective_bounds_from_workbook(path)
        return lower, upper, "exact-paper-bounds", repo_path(path)
    if config == "10_130":
        path = ROOT / "output" / "excel" / "nsga2_normalized_results20.xlsx"
        lower, upper = objective_bounds_from_workbook(path)
        return lower, upper, "exact-paper-bounds", repo_path(path)

    paths = audit_paths(config)
    lower, upper = objective_bounds_from_npz(paths)
    quality = "exact-paper-bounds" if config == "10_100" else "same-config-seed42-bounds"
    return lower, upper, quality, "; ".join(repo_path(path) for path in paths)


def best_q(objectives: np.ndarray, lower: np.ndarray, upper: np.ndarray) -> float:
    span = upper - lower
    if np.any(span <= 0):
        raise ValueError(f"Invalid normalization span: lower={lower}, upper={upper}")
    normalized = (objectives - lower) / span
    return float(np.min(0.5 * normalized.sum(axis=1)))


def dqn_record(config: str) -> dict[str, object]:
    path = NPZ_DIR / f"res_dqn_{config}.npz"
    if not path.exists():
        raise FileNotFoundError(path)
    with np.load(path) as data:
        objectives = np.asarray(data["F"], dtype=float)
        seeds = np.asarray(data["seeds"]).astype(int).tolist()
    if seeds != [42]:
        raise ValueError(f"Expected DQN seed [42] for {config}, found {seeds}")

    lower, upper, quality, evidence = normalization_bounds(config)
    return {
        "Config": config,
        "Method": "DQN",
        "BestQ": best_q(objectives, lower, upper),
        "Source": repo_path(path),
        "NormalizationClass": quality,
        "NormalizationEvidence": evidence,
        "ProtocolSeed": 42,
        "CostLower": float(lower[0]),
        "CostUpper": float(upper[0]),
        "DelayLower": float(lower[1]),
        "DelayUpper": float(upper[1]),
    }


def main() -> None:
    archive = pd.read_csv(ARCHIVE_CSV)
    expected = {(config, method) for config in CONFIG_ORDER for method in METHOD_ORDER[:-1]}
    actual = set(zip(archive["Config"], archive["Method"]))
    if actual != expected:
        missing = sorted(expected - actual)
        extra = sorted(actual - expected)
        raise ValueError(f"Archive coverage mismatch. Missing={missing}; extra={extra}")

    records: list[dict[str, object]] = []
    for row in archive.itertuples(index=False):
        records.append(
            {
                "Config": row.Config,
                "Method": row.Method,
                "BestQ": float(row.BestQ),
                "Source": row.Source,
                "NormalizationClass": "original-paper-archive",
                "NormalizationEvidence": repo_path(ARCHIVE_CSV),
                "ProtocolSeed": 42,
                "CostLower": None,
                "CostUpper": None,
                "DelayLower": None,
                "DelayUpper": None,
            }
        )
    records.extend(dqn_record(config) for config in CONFIG_ORDER)

    table = pd.DataFrame.from_records(records)
    table["Config"] = pd.Categorical(table["Config"], CONFIG_ORDER, ordered=True)
    table["Method"] = pd.Categorical(table["Method"], METHOD_ORDER, ordered=True)
    table = table.sort_values(["Config", "Method"]).reset_index(drop=True)

    validation_rows = []
    improvements = []
    for config in CONFIG_ORDER:
        group = table[table["Config"] == config]
        values = {str(row.Method): float(row.BestQ) for row in group.itertuples(index=False)}
        winner = min(values, key=values.get)
        best_alternative = min(values[method] for method in ["NS-P", "GCP", "GDP"])
        improvement = 100.0 * (best_alternative - values["PSP"]) / best_alternative
        improvements.append(improvement)
        validation_rows.append(
            {
                "Config": config,
                "Winner": winner,
                "PSP_BestQ": values["PSP"],
                "Best_Non_PSP_Evolutionary_Q": best_alternative,
                "PSP_Reduction_Pct": improvement,
                "DQN_BestQ": values["DQN"],
                "PSP_Is_Lowest": winner == "PSP",
            }
        )
    validation = pd.DataFrame(validation_rows)
    if not bool(validation["PSP_Is_Lowest"].all()):
        raise AssertionError("PSP is not the lowest-Q method in every archived paper configuration")

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    table.to_csv(OUT_CSV, index=False, float_format="%.10f")
    validation.to_csv(OUT_VALIDATION, index=False, float_format="%.10f")

    value_map = {
        config: {
            method: float(
                table[(table["Config"] == config) & (table["Method"] == method)]["BestQ"].iloc[0]
            )
            for method in METHOD_ORDER
        }
        for config in CONFIG_ORDER
    }
    payload = {
        "methods": METHOD_ORDER,
        "config_order": CONFIG_ORDER,
        "values": value_map,
        "records": json.loads(
            table.astype({"Config": "object", "Method": "object"}).to_json(orient="records")
        ),
        "validation": validation.to_dict(orient="records"),
        "summary": {
            "psp_lowest_all_configurations": True,
            "configuration_count": len(CONFIG_ORDER),
            "minimum_reduction_pct": float(min(improvements)),
            "maximum_reduction_pct": float(max(improvements)),
            "mean_reduction_pct": float(np.mean(improvements)),
            "seed": 42,
        },
    }
    OUT_JSON.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8")

    print(table[["Config", "Method", "BestQ", "NormalizationClass"]].to_string(index=False))
    print("\nValidation")
    print(validation.to_string(index=False))
    print(f"\nSaved {OUT_CSV}")
    print(f"Saved {OUT_JSON}")
    print(f"Saved {OUT_VALIDATION}")


if __name__ == "__main__":
    main()
